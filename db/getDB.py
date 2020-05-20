import pymysql
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import ExcelTitle


def getConnection(db):
    return pymysql.connect(host='localhost', port=3306, user='root', password='123456', db=db)

def getColumnFormMysql(connection):
    cur = connection.cursor(cursor=pymysql.cursors.DictCursor)
    cur.execute('show tables')
    return cur.fetchall()


def getFieldFromDB(table, connection):
    cur = connection.cursor(cursor=pymysql.cursors.DictCursor)
    cur.execute('show full columns from ' + table)
    return cur.fetchall()


def savInfoToExcel(ws, fieldInfo):
    row_idx = 5
    for iterating_var in fieldInfo:
        idx = 1
        ws.cell(row=row_idx, column=idx).value = str(row_idx-4)
        ExcelTitle.writeBorder(ws.cell(row=row_idx, column=idx))
        idx += 1
        for key in iterating_var:
            if key.lower() == 'PRIVILEGES'.lower():
                continue
            if key == 'Type':
                p1 = re.findall(r'\(.*?\)', iterating_var[key])
                type_len = iterating_var[key].find('(') if iterating_var[key].find('(') != -1 else len(iterating_var[key])
                ws.cell(row=row_idx, column=idx).value = iterating_var[key][0:type_len]
                ExcelTitle.writeBorder(ws.cell(row=row_idx, column=idx))
                idx += 1
                if len(p1) == 1:
                    ws.cell(row=row_idx, column=idx).value = p1[0][1:len(p1[0]) -1]
            else:
                ws.cell(row=row_idx, column=idx).value = iterating_var[key]
            ExcelTitle.writeBorder(ws.cell(row=row_idx, column=idx))
            idx += 1
        row_idx += 1
    pass


def resize(ws, fieldInfo):
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 10
    ws.column_dimensions[get_column_letter(6)].width = 10
    for i in range(1,fieldInfo):
        ws.column_dimensions[get_column_letter(i+1)].width = 25

if __name__ == "__main__":
    dbList = ['test']
    for db in dbList:
        connection = getConnection(db);
        tables = getColumnFormMysql(connection);
        wb = Workbook()
        idx = 2
        for iterating_var in tables:
            key = ''
            for keyI in iterating_var:
                key = keyI
                pass
            fieldInfo = getFieldFromDB(iterating_var[key], connection)
            ws1 = wb.create_sheet(iterating_var[key])
            savInfoToExcel(ws1, fieldInfo)
            ExcelTitle.editExcelTitle(ws1,iterating_var[key],idx)
            resize(ws1,11)
            ExcelTitle.createLink(wb,iterating_var[key],idx)
            idx += 1
        wb.save('/Users/admin/Desktop/person/' + db + '.xlsx')
