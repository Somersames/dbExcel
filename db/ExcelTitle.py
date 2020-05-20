
from openpyxl.styles import PatternFill, Border, Side, Font, colors, Alignment
from openpyxl.utils import get_column_letter


def editExcelTitle(ws, table, idx):
    title = ['序号', '字段定义', '字段类型', '精度', '字符集', '是否可为空', '索引', '默认值', '额外备注', '注释']
    idx = 1
    fill = PatternFill("solid", fgColor="fcd5b4")
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    for t in title:
        ws.cell(row=4, column=idx).value = t
        d = ws.cell(row=4, column=idx)
        d.fill = fill
        d.border = border
        idx += 1

    ws.merge_cells('A2:A4')
    top_left_cell = ws['A2']
    top_left_cell.value = 'No'
    top_left_cell.fill = fill
    top_left_cell.border = border
    ws['B2'] = '表中文名称'
    ws['B3'] = '表英文名称'
    ws['C3'] = table
    link = '#Sheet!A' + str(idx)
    ws['A1'].value = '=HYPERLINK("%s", "%s")' % (link, '返回')

def createLink(wb,sheet,index):
    sheettitle = ['No','表名','表定义','描述']
    fill = PatternFill("solid", fgColor="fcd5b4")
    green = PatternFill("solid", fgColor="ffcc66")
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    ws = wb.get_sheet_by_name("Sheet")
    link = '#' + sheet + '!A1'
    for key in range(1,len(sheettitle)):
        ws.cell(row=1, column=key).value = sheettitle[key-1]
        ws.cell(row=1, column=key).fill = fill
        ws.column_dimensions[get_column_letter(key + 1)].width = 40
        ws.cell(row=1, column=key).border = border
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws['B' + str(index)].value = '=HYPERLINK("%s", "%s")' % (link, sheet)
    ws['B' + str(index)].font = Font(u='single', color=colors.BLUE)
    ws['A' + str(index)].value = index -1
    aligmentCenter = Alignment(horizontal='center', vertical ='center')
    ws['A' + str(index)].alignment = aligmentCenter
    ws['A' + str(index)].fill = green
    ws['A' + str(index)].border = border

def writeBorder(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    ws.border = border
